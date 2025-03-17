import streamlit as st
import pandas as pd
import yfinance as yf
import numpy as np
from datetime import datetime
from fuzzywuzzy import fuzz, process
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, NamedStyle, Alignment
from openpyxl.utils import get_column_letter
from io import BytesIO
import plotly.graph_objects as go
import logging
import time
import re

# Configure logging
logging.basicConfig(level=logging.INFO, 
                   format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# Set Streamlit page configuration
st.set_page_config(
    page_title="Financial Analysis Suite",
    page_icon="📊",
    layout="wide"
)

class FinancialDataAnalyzer:
    def __init__(self):
        self.statement_types = {
            'Balance Sheet': 'balance_sheet',
            'Income Statement': 'income_stmt',
            'Cash Flow Statement': 'cash_flow'
        }

    def get_company_name(self, ticker):
        """Get full company name from ticker"""
        try:
            company = yf.Ticker(ticker)
            return company.info.get('longName', ticker)
        except Exception as e:
            logger.warning(f"Error getting company name for {ticker}: {str(e)}")
            return ticker

    def find_best_comparison_date(self, companies_data, statement_type):
        """
        Find the most appropriate comparison date across companies
        """
        all_dates = set()
        company_dates = {}
        
        for company, data in companies_data.items():
            if statement_type in data['data']:
                df = data['data'][statement_type]
                if df is not None and not df.empty:
                    dates = df.columns
                    company_dates[company] = dates
                    all_dates.update(dates)
                else:
                    logger.warning(f"No data available for {statement_type} of {company}")
        
        all_dates = sorted(list(all_dates), reverse=True)
        
        if not all_dates:
            logger.warning(f"No dates available for {statement_type} across companies")
            return {}
            
        common_date = None
        for date in all_dates:
            companies_with_date = sum(1 for dates in company_dates.values() if date in dates)
            if companies_with_date > 1:
                common_date = date
                break
        
        best_dates = {}
        for company, dates in company_dates.items():
            if common_date and common_date in dates:
                best_dates[company] = common_date
            elif len(dates) > 0:
                best_dates[company] = dates[0]
            else:
                best_dates[company] = None
                logger.warning(f"No valid dates for {company} in {statement_type}")
                
        return best_dates

    def get_base_value(self, df, statement_type, date_col):
        """
        Safely get the base value for common size analysis,
        with fallbacks in case primary metrics aren't available
        """
        if statement_type == 'Balance Sheet':
            if 'Total Assets' in df.index:
                return df.loc['Total Assets', date_col]
            elif 'Total assets' in df.index:
                return df.loc['Total assets', date_col]
            else:
                assets_rows = [i for i in df.index if 'asset' in str(i).lower()]
                if assets_rows:
                    return df.loc[assets_rows[0], date_col]
                logger.warning(f"No base value found for {statement_type}, using 1 as fallback")
                return 1
        elif statement_type == 'Income Statement':
            if 'Total Revenue' in df.index:
                return df.loc['Total Revenue', date_col]
            elif 'Revenue' in df.index:
                return df.loc['Revenue', date_col]
            else:
                revenue_rows = [i for i in df.index if 'revenue' in str(i).lower() or 'sales' in str(i).lower()]
                if revenue_rows:
                    return df.loc[revenue_rows[0], date_col]
                logger.warning(f"No base value found for {statement_type}, using 1 as fallback")
                return 1
        else:  # Cash Flow Statement
            if 'Operating Cash Flow' in df.index:
                return df.loc['Operating Cash Flow', date_col]
            elif 'Cash from Operations' in df.index:
                return df.loc['Cash from Operations', date_col]
            else:
                cf_rows = [i for i in df.index if 'operating' in str(i).lower() and 'cash' in str(i).lower()]
                if cf_rows:
                    return df.loc[cf_rows[0], date_col]
                logger.warning(f"No base value found for {statement_type}, using 1 as fallback")
                return 1

    def create_comparative_statements(self, companies_data, analysis_type="Original"):
        """
        Create comparative financial statements with companies side by side
        """
        comparative_statements = {}
        
        for statement_name, statement_key in self.statement_types.items():
            best_dates = self.find_best_comparison_date(companies_data, statement_name)
            if not best_dates:
                logger.warning(f"No comparison dates found for {statement_name}")
                continue
                
            comparative_data = {}
            for company, data in companies_data.items():
                if statement_name in data['data'] and company in best_dates:
                    df = data['data'][statement_name]
                    if df is not None and not df.empty and best_dates[company] is not None:
                        date_col = best_dates[company]
                        try:
                            company_data = df[date_col]
                            if analysis_type == "Common Size":
                                base = self.get_base_value(df, statement_name, date_col)
                                if base != 0:
                                    company_data = (company_data / abs(base) * 100).round(2)
                                else:
                                    company_data = company_data.copy()
                                    logger.warning(f"Base value is zero for {company} in {statement_name}, skipping common size transformation")
                            date_str = pd.to_datetime(date_col).strftime('%Y-%m-%d')
                            comparative_data[f"{company} ({date_str})"] = company_data
                        except Exception as e:
                            logger.error(f"Error processing {company} data for {statement_name}: {str(e)}")
                    else:
                        logger.warning(f"Data for {company} in {statement_name} is empty or unavailable")
            
            if comparative_data:
                comparative_statements[statement_name] = pd.DataFrame(comparative_data)
            else:
                logger.warning(f"No comparative data generated for {statement_name}")
        
        return comparative_statements

    def fetch_financial_data(self, tickers, statements, max_retries=3):
        """Fetch financial data for multiple companies with retries"""
        results = {}
        for ticker in tickers:
            ticker = ticker.strip().upper()
            try:
                company = yf.Ticker(ticker)
                company_name = self.get_company_name(ticker)
                results[ticker] = {
                    'name': company_name,
                    'data': {}
                }
                
                for statement in statements:
                    for attempt in range(max_retries):
                        try:
                            if statement == 'Balance Sheet':
                                df = company.balance_sheet
                            elif statement == 'Income Statement':
                                df = company.income_stmt
                            else:  # Cash Flow Statement
                                df = company.cash_flow
                            
                            if df is not None and not df.empty:
                                df.columns = pd.to_datetime(df.columns)
                                results[ticker]['data'][statement] = df
                                logger.info(f"Successfully fetched {statement} for {ticker} (Attempt {attempt + 1})")
                                break
                            else:
                                logger.warning(f"Empty {statement} data for {ticker} (Attempt {attempt + 1})")
                                if attempt == max_retries - 1:
                                    st.warning(f"Could not fetch {statement} for {ticker}: Data is empty")
                        except Exception as e:
                            logger.error(f"Error fetching {statement} for {ticker} (Attempt {attempt + 1}): {str(e)}")
                            if attempt == max_retries - 1:
                                st.warning(f"Could not fetch {statement} for {ticker}: {str(e)}")
                            time.sleep(2 ** attempt)  # Exponential backoff
                        
            except Exception as e:
                error_msg = f"Error fetching data for {ticker}: {str(e)}"
                logger.error(error_msg)
                st.error(error_msg)
                
        return results

    def export_to_excel(self, data, analysis_type="Original"):
        """Export to Excel with comparative statements"""
        output = BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            sheet_written = False
            
            comparative_statements = self.create_comparative_statements(data, analysis_type)
            
            for statement_type, comp_df in comparative_statements.items():
                sheet_name = f"Comparative {statement_type}"[:31]
                comp_df.to_excel(writer, sheet_name=sheet_name)
                sheet_written = True
                
                worksheet = writer.sheets[sheet_name]
                for idx, col in enumerate(comp_df.columns):
                    max_length = max(
                        comp_df.index.astype(str).map(len).max(),
                        len(str(col)),
                        comp_df[col].astype(str).map(len).max()
                    )
                    worksheet.column_dimensions[get_column_letter(idx + 2)].width = min(max_length + 2, 50)
                worksheet.freeze_panes = 'B2'
            
            for ticker, company_data in data.items():
                company_name = company_data['name']
                for statement_type, df in company_data['data'].items():
                    if df is not None and not df.empty:
                        sheet_name = f"{company_name[:20]} {statement_type[:10]}"
                        sheet_name = sheet_name.replace('/', '-').replace(':', '-')[:31]
                        export_df = df.copy()
                        if analysis_type == "Common Size":
                            try:
                                for date_col in export_df.columns:
                                    base = self.get_base_value(export_df, statement_type, date_col)
                                    if base != 0:
                                        export_df[date_col] = (export_df[date_col] / abs(base) * 100).round(2)
                            except Exception as e:
                                logger.error(f"Error in common size calculation for {ticker} {statement_type}: {str(e)}")
                        try:
                            export_df.to_excel(writer, sheet_name=sheet_name)
                            sheet_written = True
                            worksheet = writer.sheets[sheet_name]
                            for idx, col in enumerate(export_df.columns):
                                max_length = max(
                                    export_df.index.astype(str).map(len).max(),
                                    len(str(col)),
                                    export_df[col].astype(str).map(len).max()
                                )
                                worksheet.column_dimensions[get_column_letter(idx + 2)].width = min(max_length + 2, 50)
                            worksheet.freeze_panes = 'B2'
                        except Exception as e:
                            logger.error(f"Error writing sheet {sheet_name}: {str(e)}")
            
            if not sheet_written:
                logger.warning("No sheets were written to Excel. Adding a default sheet.")
                pd.DataFrame({"Message": ["No data available for the selected companies and statements."]}).to_excel(
                    writer, sheet_name="No Data", index=False
                )

        return output.getvalue()


class FuzzyMatcher:
    def __init__(self):
        self.default_threshold = 80

    def preprocess_text(self, text):
        if pd.isna(text):
            return ""
        return str(text).lower().strip()

    def perform_fuzzy_match(self, df1, df2, col1, col2, threshold=None):
        if threshold is None:
            threshold = self.default_threshold

        matches = []
        unmatched = []

        df1_copy = df1.copy()
        df2_copy = df2.copy()

        df1_copy[col1] = df1_copy[col1].apply(self.preprocess_text)
        df2_copy[col2] = df2_copy[col2].apply(self.preprocess_text)

        df1_copy = df1_copy[df1_copy[col1] != ""].reset_index(drop=True)
        df2_copy = df2_copy[df2_copy[col2] != ""].reset_index(drop=True)

        if df1_copy.empty or df2_copy.empty:
            st.warning("One or both datasets are empty after preprocessing.")
            return pd.DataFrame(), pd.DataFrame()

        lookup_dict = dict(zip(df2_copy[col2], range(len(df2_copy))))
        
        if not lookup_dict:
            st.warning("No valid entries found in the second dataset for matching.")
            return pd.DataFrame(), pd.DataFrame()

        for idx1, row1 in df1_copy.iterrows():
            try:
                best_match = process.extractOne(row1[col1], lookup_dict.keys())
                if best_match and best_match[1] >= threshold:
                    match_idx = lookup_dict[best_match[0]]
                    matches.append({
                        'Source': row1[col1],
                        'Match': best_match[0],
                        'Similarity': best_match[1],
                        'Source_Row': idx1,
                        'Match_Row': match_idx
                    })
                else:
                    unmatched.append({
                        'Source': row1[col1],
                        'Source_Row': idx1
                    })
            except Exception as e:
                logger.error(f"Error matching row {idx1}: {str(e)}")
                unmatched.append({
                    'Source': row1[col1],
                    'Source_Row': idx1,
                    'Error': str(e)
                })

        matches_df = pd.DataFrame(matches) if matches else pd.DataFrame()
        unmatched_df = pd.DataFrame(unmatched) if unmatched else pd.DataFrame()

        return matches_df, unmatched_df

    def export_matches_to_excel(self, matches_df, unmatched_df, df1, df2, col1, col2):
        output = BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            if not matches_df.empty:
                detailed_matches = []
                for _, match in matches_df.iterrows():
                    try:
                        source_row = int(match['Source_Row'])
                        match_row = int(match['Match_Row'])
                        if 0 <= source_row < len(df1) and 0 <= match_row < len(df2):
                            source_data = df1.iloc[source_row].to_dict()
                            match_data = df2.iloc[match_row].to_dict()
                            row_data = {
                                'Similarity': match['Similarity'],
                                **{f'Source_{k}': v for k, v in source_data.items()},
                                **{f'Match_{k}': v for k, v in match_data.items()}
                            }
                            detailed_matches.append(row_data)
                    except Exception as e:
                        logger.error(f"Error processing match detail: {str(e)}")
                if detailed_matches:
                    detailed_matches_df = pd.DataFrame(detailed_matches)
                    detailed_matches_df.to_excel(writer, sheet_name='Detailed Matches', index=False)

            if not matches_df.empty:
                matches_df.to_excel(writer, sheet_name='Matches Summary', index=False)
            else:
                pd.DataFrame(columns=['Source', 'Match', 'Similarity', 'Source_Row', 'Match_Row']).to_excel(
                    writer, sheet_name='Matches Summary', index=False
                )
            
            if not unmatched_df.empty:
                unmatched_df.to_excel(writer, sheet_name='Unmatched Items', index=False)
            else:
                pd.DataFrame(columns=['Source', 'Source_Row']).to_excel(
                    writer, sheet_name='Unmatched Items', index=False
                )

            df1_export = df1.copy()
            df2_export = df2.copy()
            df1_export['Row Number'] = range(1, len(df1_export) + 1)
            df2_export['Row Number'] = range(1, len(df2_export) + 1)
            df1_export.to_excel(writer, sheet_name='Original Dataset 1', index=False)
            df2_export.to_excel(writer, sheet_name='Original Dataset 2', index=False)

        return output.getvalue()


def read_file(file):
    """Safely read uploaded file with proper error handling"""
    try:
        if file.name.lower().endswith('.csv'):
            return pd.read_csv(file)
        elif file.name.lower().endswith(('.xls', '.xlsx')):
            return pd.read_excel(file)
        else:
            st.error(f"Unsupported file format: {file.name}")
            return None
    except Exception as e:
        st.error(f"Error reading file {file.name}: {str(e)}")
        return None


def main():
    st.title("Financial Analysis Suite")
    st.write("Analyze financial statements and perform fuzzy matching on datasets")

    tab1, tab2 = st.tabs(["Financial Statement Analysis", "Fuzzy Matching"])

    with tab1:
        st.header("Financial Statement Analysis")
        
        analyzer = FinancialDataAnalyzer()
        
        tickers_input = st.text_area(
            "Enter ticker symbols (separated by semicolons or commas)",
            "AAPL;MSFT;GOOGL"
        )
        
        # Validate and parse tickers, accepting both commas and semicolons
        tickers = [t.strip().upper() for t in re.split('[,;]', tickers_input) if t.strip()]
        if not tickers:
            st.error("Please enter at least one valid ticker symbol separated by semicolons or commas (e.g., AAPL;MSFT;GOOGL or AAPL,MSFT,GOOGL)")
            return
        
        # Display the parsed tickers for user confirmation
        st.write("Selected tickers:", ", ".join(tickers))
        
        selected_statements = st.multiselect(
            "Select Financial Statements",
            ["Balance Sheet", "Income Statement", "Cash Flow Statement"],
            default=["Income Statement"]
        )
        
        analysis_type = st.selectbox(
            "Select Analysis Type",
            ["Original", "Common Size"]
        )
        
        if st.button("Fetch and Analyze"):
            if not selected_statements:
                st.error("Please select at least one statement type")
                return
                
            with st.spinner("Fetching financial data..."):
                try:
                    results = analyzer.fetch_financial_data(tickers, selected_statements)
                    
                    # Check if any data was fetched
                    has_data = False
                    failed_tickers = []
                    for ticker, company_data in results.items():
                        if company_data['data']:
                            has_data = True
                        else:
                            failed_tickers.append(ticker)
                    
                    if not has_data:
                        st.error("No financial data could be fetched for the selected companies and statements.")
                        if failed_tickers:
                            st.warning(f"Failed to fetch data for tickers: {', '.join(failed_tickers)}")
                        return
                    
                    # Display comparative statements
                    comparative_statements = analyzer.create_comparative_statements(results, analysis_type)
                    
                    if not comparative_statements:
                        st.warning("Could not create comparative statements. Data may be unavailable.")
                    else:
                        for statement_type, comp_df in comparative_statements.items():
                            st.subheader(f"Comparative {statement_type}")
                            st.dataframe(comp_df)
                    
                    # Create Excel export
                    excel_data = analyzer.export_to_excel(results, analysis_type)
                    
                    st.download_button(
                        label="Download Complete Analysis",
                        data=excel_data,
                        file_name=f"financial_analysis_{datetime.now().strftime('%Y%m%d')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                except Exception as e:
                    st.error(f"An error occurred during analysis: {str(e)}")
                    logger.exception("Error in financial analysis")

    with tab2:
        st.header("Fuzzy Matching Analysis")
        
        matcher = FuzzyMatcher()

        upload_col1, upload_col2 = st.columns(2)
        
        with upload_col1:
            st.subheader("First Dataset")
            file1 = st.file_uploader("Upload first CSV/Excel file", type=['csv', 'xlsx', 'xls'], key="file1")
            
        with upload_col2:
            st.subheader("Second Dataset")
            file2 = st.file_uploader("Upload second CSV/Excel file", type=['csv', 'xlsx', 'xls'], key="file2")

        if file1 and file2:
            df1 = read_file(file1)
            df2 = read_file(file2)

            if df1 is not None and df2 is not None:
                if df1.empty:
                    st.error("First dataset is empty")
                elif df2.empty:
                    st.error("Second dataset is empty")
                else:
                    select_col1, select_col2 = st.columns(2)
                    
                    with select_col1:
                        match_col1 = st.selectbox("Select matching column from first dataset", df1.columns)
                        
                    with select_col2:
                        match_col2 = st.selectbox("Select matching column from second dataset", df2.columns)

                    threshold = st.slider("Matching threshold (%)", 0, 100, 80)

                    if st.button("Perform Matching"):
                        try:
                            with st.spinner("Performing fuzzy matching..."):
                                matches_df, unmatched_df = matcher.perform_fuzzy_match(
                                    df1, df2, match_col1, match_col2, threshold
                                )

                                st.subheader("Matching Results")
                                if not matches_df.empty:
                                    st.write(f"Found {len(matches_df)} matches")
                                    st.dataframe(matches_df)
                                else:
                                    st.info("No matches found with the current threshold. Try lowering the threshold.")
                                
                                excel_data = matcher.export_matches_to_excel(
                                    matches_df, unmatched_df, df1, df2, match_col1, match_col2
                                )
                                
                                st.download_button(
                                    label="Download Matching Results (Excel)",
                                    data=excel_data,
                                    file_name=f"fuzzy_matching_results_{datetime.now().strftime('%Y%m%d')}.xlsx",
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                                )
                                
                                if not unmatched_df.empty:
                                    st.subheader("Unmatched Items")
                                    st.write(f"Found {len(unmatched_df)} unmatched items")
                                    st.dataframe(unmatched_df)
                        except Exception as e:
                            st.error(f"Error during matching: {str(e)}")
                            logger.exception("Error in fuzzy matching")
                 
if __name__ == "__main__":
    main()
